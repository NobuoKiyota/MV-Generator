import os
import json
import librosa
import google.generativeai as genai
from dotenv import load_dotenv
import re
import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import mido

load_dotenv()

def parse_tagged_lyrics(raw_lyrics):
    """
    [Intro], [サビ], [A-Melody] などのタグで区切られた歌詞をパースして
    セクションごとのリストを作成する
    """
    if not raw_lyrics:
        return []
        
    pattern = r'\[([^\]]+)\]'
    segments = re.split(pattern, raw_lyrics)
    
    parsed = []
    
    # 最初がタグから始まらない場合のテキスト
    first_text = segments[0].strip()
    if first_text:
        parsed.append({"section": "Intro", "lyrics": first_text})
        
    for i in range(1, len(segments), 2):
        section_name = segments[i].strip()
        lyrics_text = segments[i+1].strip() if i+1 < len(segments) else ""
        if lyrics_text or section_name:
            parsed.append({"section": section_name, "lyrics": lyrics_text})
            
    return parsed

def get_lyrics_lines_with_sections(raw_lyrics):
    """
    歌詞をセクションごとにパースし、空行を除外した『行単位』のリストにする
    例: [{"section": "Verse 1", "line": "届かない"}, {"section": "Verse 1", "line": "響かない"}]
    """
    sections = parse_tagged_lyrics(raw_lyrics)
    line_list = []
    
    for sec in sections:
        sec_name = sec["section"]
        lyrics_text = sec["lyrics"]
        lines = [l.strip() for l in lyrics_text.split("\n") if l.strip()]
        
        # 歌詞が空の場合はダミー行を1つ置く (インストセクションなど)
        if not lines:
            line_list.append({"section": sec_name, "line": ""})
        else:
            for l in lines:
                line_list.append({"section": sec_name, "line": l})
                
    return line_list

def align_lyrics_with_midi(midi_path, raw_lyrics, duration):
    """
    MIDIファイルのメロディ音符（Note On）から発音区間（フレーズ）を抽出し、
    入力された歌詞（行単位）とミリ秒単位で完全に同期させた初期タイムライン枠を作成する。
    """
    try:
        mid = mido.MidiFile(midi_path)
    except Exception as e:
        print(f"Error loading MIDI file: {e}")
        return None

    # 1. MIDI全イベントを絶対秒数に解決してノートオンを抽出
    notes = [] # (start_time, end_time, note_number)
    active_notes = {}
    current_time = 0.0
    
    # マーカー情報（もしあればセクション名アライメントに使用）
    markers = [] # (time, label)
    
    for msg in mid:
        current_time += msg.time
        if msg.type == 'marker':
            markers.append((current_time, msg.text.strip()))
        elif msg.type == 'note_on' and msg.velocity > 0:
            active_notes[msg.note] = current_time
        elif (msg.type == 'note_off') or (msg.type == 'note_on' and msg.velocity == 0):
            if msg.note in active_notes:
                start_time = active_notes.pop(msg.note)
                end_time = current_time
                # 極端に短いノイズ音符（0.05秒以下）は無視
                if end_time - start_time > 0.05:
                    notes.append((start_time, end_time, msg.note))

    # ノートを時系列ソート
    notes.sort(key=lambda x: x[0])
    
    # 2. 発音タイミングを「休符（無音区間）」で区切り、フレーズ（グループ）を抽出
    # ノート同士の間が 1.5 秒以上空いた場合、別フレーズ（カット）とする
    phrase_groups = []
    if notes:
        current_group = [notes[0]]
        for note in notes[1:]:
            last_note = current_group[-1]
            rest_duration = note[0] - last_note[1]
            if rest_duration > 1.5:
                phrase_groups.append(current_group)
                current_group = [note]
            else:
                current_group.append(note)
        phrase_groups.append(current_group)

    # 各グループの開始・終了時間（秒）を計算
    midi_phrases = []
    for gp in phrase_groups:
        start_time = gp[0][0]
        end_time = gp[-1][1]
        # バッファ（歌い出しの前後に0.2秒の余白を設ける）
        start_time = max(0.0, start_time - 0.2)
        end_time = min(duration, end_time + 0.2)
        midi_phrases.append({"start_time": round(start_time, 3), "end_time": round(end_time, 3)})

    # 3. 歌詞テキストをセクション・行単位に分解
    lyrics_lines = get_lyrics_lines_with_sections(raw_lyrics)

    # 4. MIDIのフレーズ枠と歌詞行を時系列でマッピング (アライメント)
    timeline_slots = []
    
    # 前奏区間の追加 (最初の音符が始まるまで)
    first_singing_time = midi_phrases[0]["start_time"] if midi_phrases else 0.0
    if first_singing_time > 1.0:
        # マーカーからIntroセクションのラベル名があれば使う
        intro_label = "Intro"
        if markers:
            # 最初のマーカーが歌い出しより前ならそのテキストを使用
            if markers[0][0] < first_singing_time:
                intro_label = markers[0][1]
        timeline_slots.append({
            "section": intro_label,
            "start_time": 0.0,
            "end_time": first_singing_time,
            "lyrics": ""
        })

    # フレーズごとの割り当て
    num_midi_phrases = len(midi_phrases)
    num_lyrics_lines = len(lyrics_lines)
    
    if num_lyrics_lines == 0:
        # 歌詞がない場合はMIDIのフレーズをそのままインストカットにする
        for i, phrase in enumerate(midi_phrases):
            timeline_slots.append({
                "section": f"Instrumental {i+1}",
                "start_time": phrase["start_time"],
                "end_time": phrase["end_time"],
                "lyrics": ""
            })
    else:
        # 歌詞行とMIDIフレーズの同期マッピング
        # 基本的に1対1でマッピングするが、数が合わない場合は補正
        for idx in range(max(num_midi_phrases, num_lyrics_lines)):
            if idx < num_midi_phrases and idx < num_lyrics_lines:
                # 1対1マッチ
                phrase = midi_phrases[idx]
                lyr_info = lyrics_lines[idx]
                timeline_slots.append({
                    "section": lyr_info["section"],
                    "start_time": phrase["start_time"],
                    "end_time": phrase["end_time"],
                    "lyrics": lyr_info["line"]
                })
            elif idx < num_midi_phrases:
                # 歌詞が足りない場合: 残りはインストにする
                phrase = midi_phrases[idx]
                timeline_slots.append({
                    "section": "Instrumental/Solo",
                    "start_time": phrase["start_time"],
                    "end_time": phrase["end_time"],
                    "lyrics": ""
                })
            elif idx < num_lyrics_lines:
                # MIDI音符が足りない場合: 最後の音符以降に順次4秒間隔などで割り当てる
                last_end = timeline_slots[-1]["end_time"] if timeline_slots else 0.0
                lyr_info = lyrics_lines[idx]
                start = round(last_end, 3)
                end = round(min(duration, start + 4.0), 3)
                if start < duration:
                    timeline_slots.append({
                        "section": lyr_info["section"],
                        "start_time": start,
                        "end_time": end,
                        "lyrics": lyr_info["line"]
                    })

    # カット間の「隙間」を埋める (小節グリッドまたはシームレス接続)
    # カットとカットの間の無音区間を自動で繋ぐ
    seamless_slots = []
    for i in range(len(timeline_slots)):
        current_slot = timeline_slots[i]
        
        # 最初のスロットが0.0秒から始まっていない場合
        if i == 0 and current_slot["start_time"] > 0.0:
            seamless_slots.append({
                "section": "Intro",
                "start_time": 0.0,
                "end_time": current_slot["start_time"],
                "lyrics": ""
            })
            
        seamless_slots.append(current_slot)
        
        # 次のスロットとの間に隙間があるかチェック
        if i < len(timeline_slots) - 1:
            next_slot = timeline_slots[i+1]
            gap = next_slot["start_time"] - current_slot["end_time"]
            if gap > 0.5: # 0.5秒以上の隙間があれば、間奏・無音カットを作る
                seamless_slots.append({
                    "section": "Interlude",
                    "start_time": current_slot["end_time"],
                    "end_time": next_slot["start_time"],
                    "lyrics": ""
                })
            elif gap > 0.0: # 極小の隙間なら前のカットを後ろに伸ばして繋ぐ
                current_slot["end_time"] = next_slot["start_time"]

    # 最後のカットを曲の終端 (duration) まで引き伸ばす
    if seamless_slots and seamless_slots[-1]["end_time"] < duration:
        last_slot = seamless_slots[-1]
        if last_slot["lyrics"] == "":
            last_slot["end_time"] = duration
        else:
            seamless_slots.append({
                "section": "Outro",
                "start_time": last_slot["end_time"],
                "end_time": duration,
                "lyrics": ""
            })

    # 重複や逆転などのバグをクリーンアップ
    valid_slots = []
    for slot in seamless_slots:
        if slot["end_time"] > slot["start_time"]:
            valid_slots.append(slot)
            
    return valid_slots

def save_project(filepath, data):
    """プロジェクトデータをJSONファイルとして保存する"""
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving project: {e}")
        return False

def load_project(filepath):
    """プロジェクトデータをJSONファイルから読み込む"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading project: {e}")
        return None

def analyze_audio(audio_path):
    """音声ファイルを解析してBPMと曲の長さを返す"""
    try:
        y, sr = librosa.load(audio_path, sr=None)
        duration = librosa.get_duration(y=y, sr=sr)
        
        # BPMの推定
        tempo, beat_frames = librosa.beat.beat_track(y=y, sr=sr)
        
        # テンポ（BPM）はfloatまたはarrayで返るため数値に変換
        if isinstance(tempo, (list, tuple, bytes, dict)) or hasattr(tempo, '__iter__'):
            bpm = float(tempo[0]) if len(tempo) > 0 else 120.0
        else:
            bpm = float(tempo)
            
        return {
            "bpm": round(bpm, 2),
            "duration": round(duration, 2)
        }
    except Exception as e:
        print(f"Error analyzing audio: {e}")
        return None

def generate_initial_timeline(lyrics, bpm, duration, core_concept="", global_rules="", section_rules="", keywords="", api_key=None,
                              first_bar_offset=0.0, time_signature="4/4", split_unit="4小節ごと",
                              visual_style="特になし", aspect_ratio="16:9", audio_path=None, midi_path=None):
    """Gemini APIを呼び出し、音源・歌詞・テンポ・MIDIから、世界観、演出ルール、および同期したタイムライン・プロンプトを全自動生成する"""
    if api_key:
        genai.configure(api_key=api_key)
    elif os.environ.get("GEMINI_API_KEY"):
        genai.configure(api_key=os.environ.get("GEMINI_API_KEY"))
    else:
        raise ValueError("Gemini API key is not configured.")

    model = genai.GenerativeModel('gemini-2.5-flash')
    
    # 1. MIDIファイルがある場合は、ローカルの超精密アライメントを実行
    timeline_slots = None
    if midi_path and os.path.exists(midi_path):
        print(f"[INFO] MIDI alignment mode activated: {midi_path}")
        timeline_slots = align_lyrics_with_midi(midi_path, lyrics, duration)
        
    # 2. MIDIがない場合は、歌詞を単純パースしたリストを作る
    if timeline_slots is None:
        print("[INFO] Fallback to standard text timeline mode.")
        structured_lyrics = parse_tagged_lyrics(lyrics)
        lyrics_json_str = json.dumps(structured_lyrics, indent=2, ensure_ascii=False)
    else:
        lyrics_json_str = json.dumps(timeline_slots, indent=2, ensure_ascii=False)
    
    # 拍子から1小節あたりの拍数を取得
    try:
        beats_per_bar = int(time_signature.split('/')[0])
    except Exception:
        beats_per_bar = 4
        
    beat_duration = 60.0 / bpm
    bar_duration = beat_duration * beats_per_bar
    
    # 世界観コンセプトなどが空の場合のAI自動生成指示
    concept_instruction = ""
    if not core_concept:
        concept_instruction = """
- 【最優先】世界観・キャラクター設定が未指定のため、あなたが楽曲の歌詞や【音声データの雰囲気・ジャンル】から、最もクリエイティブで魅力的な「世界観・コンセプト」および「キャラクター設定」を自発的に想像して自動設計してください。
- 提案する「プロジェクトのタイトル（曲名ベース）」「世界観・コンセプト」「キャラクター設定」を、出力JSONの `summary` オブジェクトに格納してください。
- すべてのカットプロンプトは、この自動設計した世界観とキャラクター描写を一貫して崩さないように強力に維持してください。
"""
    else:
        concept_instruction = f"""
- 世界観・キャラクター設定:
\"\"\"
{core_concept}
\"\"\"
上記の設定を厳格に維持し、プロンプトに反映してください。
"""

    rules_instruction = ""
    if not global_rules:
        rules_instruction = """
- 【優先】全体の演出ルールが未指定のため、あなたが楽曲の盛り上がり（イントロからサビ、アウトロ等）に合わせて、カメラワークや色彩設計、カット割りのテンポ感といった「全体の演出ルール」を自発的に設計し、出力JSONの `summary.rules` に格納してください。
"""
    else:
        rules_instruction = f"""
- 全体演出ルール / 盛り上がり制御:
\"\"\"
{global_rules}
\"\"\"
上記の演出ルールを反映させてください。
"""

    # MIDIありとなしでプロンプトを分岐
    if timeline_slots is not None:
        # MIDI同期枠がすでにある場合: 時間は完全固定。AIは演出とプロンプトだけを埋める。
        prompt = f"""
あなたはプロのミュージックビデオ(MV)監督および映像プランナーです。
入力データとして渡された【時間と歌詞が確定済みのタイムライン枠】に対し、指定の世界観・演出ルールに沿った具体的な【演出・カメラワーク（日本語）】と【画像生成プロンプト（英語）】を考えて、タイムラインを完成させてください。

【入力データ】
- 確定済みのタイムライン枠 (start_time と end_time、歌詞は完全に固定されており変更不可):
```json
{lyrics_json_str}
```
- テンポ (BPM): {bpm}
- 全体の長さ: {duration}秒
- アスペクト比: {aspect_ratio}
- 指定ビジュアルスタイル: {visual_style}

【世界観・演出設計指示】
{concept_instruction}
{rules_instruction}
{f"- セクション個別ルール: {section_rules}" if section_rules else ""}
{f"- キーワード: {keywords}" if keywords else ""}

【設計の条件】
1. 渡されたJSONの各スロットの `start_time`、`end_time`、`section`、`lyrics` は一切変更しないでください。
2. 各カットの歌詞とタイミングに完璧にフィットする、魅力的な【映像演出説明 (description)】（日本語）と、画像生成AI用の具体的かつ高品質な【画像生成プロンプト (prompt)】（英語）をあなたが考えて、JSONの空欄部分を埋めてください。
3. 指定されたアスペクト比「{aspect_ratio}」およびスタイル「{visual_style}」を強くプロンプトに反映させてください。

必ず以下のJSONスキーマに従って結果を返してください。JSON以外の文章は一切出力しないでください。

【出力フォーマット (JSON)】
{{
  "bpm": {bpm},
  "duration": {duration},
  "summary": {{
    "title": "AIが自動提案するこの楽曲のプロジェクトタイトル（日本語または英語）",
    "concept": "AIが自動設計した世界観・コンセプト設定（日本語、200〜300文字程度）",
    "characters": "AIが自動設計したキャラクター設定（日本語、100〜200文字程度）",
    "rules": "AIが自動設計した全体の演出ルール / 盛り上がり制御（日本語、100〜200文字程度）"
  }},
  "timeline": [
    // 入力と同じ順序・数で、descriptionとpromptが充填されたタイムライン
    {{
      "section": "セクション名",
      "start_time": 0.0,
      "end_time": 10.0,
      "lyrics": "歌詞の一節",
      "description": "演出・カメラワークの詳細な日本語説明",
      "prompt": "Highly detailed image generation prompt in English, matching the style and rules"
    }}
  ]
}}
"""
    else:
        # MIDIがない場合: 従来の標準モード（AIによる時間割作成、ただしBPMスナップ指示付き）
        prompt = f"""
あなたはプロのミュージックビデオ(MV)監督および映像プランナーです。
入力データと制約事項を元に、曲全体の「世界観」「キャラクター設定」「演出ルール」を自動設計し、さらにそれに完璧に同期した「タイムライン（カット割り）」を高度に設計してください。

【入力データ】
- 構造化された歌詞リスト (セクションごとの分割):
```json
{lyrics_json_str}
```
- テンポ (BPM): {bpm}
- 全体の長さ: {duration}秒
- 1小節目の開始時間 (オフセット秒): {first_bar_offset}秒
- 拍子: {time_signature} (1拍の長さ: {beat_duration:.4f}秒, 1小節の長さ: {bar_duration:.4f}秒)
- アスペクト比: {aspect_ratio}
- 指定ビジュアルスタイル: {visual_style}

【音声データの解析指示 (最重要)】
- あなたに入力として渡された音声データを実際に「聴いて」、以下の音楽的特徴を正確に解析してください。
  1. ボーカルが実際に歌い出すタイミング（何秒から声が入るか）。
  2. Aメロ、Bメロ、サビ（Chorus）などの各セクションが物理的な秒数（何秒から盛り上がるか、何秒から静かになるか）。
  3. 各歌詞テキストが歌われている正確なタイミング（アライメント）。
- 解析した物理的な歌い出しや展開のタイミングに同期させて、各カットの開始時間（start_time）および終了時間（end_time）を決定してください。

【世界観・演出設計指示】
{concept_instruction}
{rules_instruction}
{f"- セクション個別ルール: {section_rules}" if section_rules else ""}
{f"- キーワード: {keywords}" if keywords else ""}

【設計の条件】
1. 人間が指定した「構造化された歌詞リスト」の各セクションの並び順と構造を必ず遵守してください。
2. 曲全体の長さ（0.0秒から{duration}秒まで）をカバーするようにタイムラインを定義し、いくつかの「カット（Cut）」に分割してください。
3. 各カットの開始・終了時間は、実際の展開タイミングを基準としつつ、できるだけBPM {bpm} に基づく「1小節 = {bar_duration:.3f}秒」の境界線（グリッド）にスナップ（四捨五入などの補正）させて、音楽的に気持ちの良い切り替えにしてください。
4. 各カットに対して、以下を設計してください。
   - section: セクション名
   - start_time: カット開始秒数 (0.0 からスタート)
   - end_time: カット終了秒数 (最後は必ず {duration} 秒になるように)
   - lyrics: そのカットに対応する歌詞（そのタイミングで実際に歌われている歌詞フレーズ、インストの場合は空文字 ""）
   - description: 演出・カメラワークの詳細な日本語説明
   - prompt: 後の画像生成AI (Imagen 3等) で入力するための、具体的かつ高品質な英語の画像生成プロンプト。指定されたアスペクト比「{aspect_ratio}」およびスタイル「{visual_style}」を強く反映させ、世界観・キャラクター・ルールを高度にブレンドしてください。

必ず以下のJSONスキーマに従って結果を返してください。JSON以外の文章は一切出力しないでください。

【出力フォーマット (JSON)】
{{
  "bpm": {bpm},
  "duration": {duration},
  "summary": {{
    "title": "AIが自動提案するこの楽曲 of プロジェクトタイトル",
    "concept": "AIが自動設計した世界観・コンセプト設定",
    "characters": "AIが自動設計したキャラクター設定",
    "rules": "AIが自動設計した全体の演出ルール / 盛り上がり制御"
  }},
  "timeline": [
    {{
      "section": "セクション名",
      "start_time": 0.0,
      "end_time": {first_bar_offset if first_bar_offset > 0 else bar_duration},
      "lyrics": "歌詞の一節（あれば）",
      "description": "演出・カメラワークの詳細な日本語説明",
      "prompt": "Highly detailed image generation prompt in English, matching the style and rules"
    }}
  ]
}}
"""

    contents = []
    audio_file_ref = None
    
    # 音声ファイルが指定されている場合はアップロードして入力に含める
    if audio_path and os.path.exists(audio_path):
        print(f"[INFO] Uploading audio file to Gemini API: {audio_path}")
        try:
            audio_file_ref = genai.upload_file(path=audio_path)
            contents.append(audio_file_ref)
            print("[INFO] Audio upload completed.")
        except Exception as e:
            print(f"[WARNING] Failed to upload audio file to Gemini: {e}. Falling back to text-only generation.")

    contents.append(prompt)

    try:
        response = model.generate_content(
            contents,
            generation_config={
                "response_mime_type": "application/json",
                "temperature": 0.3
            }
        )
        
        result_json = json.loads(response.text)
        
        # もしMIDIアライメント結果を渡していた場合、AIが万が一秒数を変えてしまっていたら
        # ローカル側の厳密な秒数(timeline_slots)で上書き強制補正する (安全策)
        if timeline_slots is not None and result_json and "timeline" in result_json:
            for idx, slot in enumerate(result_json["timeline"]):
                if idx < len(timeline_slots):
                    slot["start_time"] = timeline_slots[idx]["start_time"]
                    slot["end_time"] = timeline_slots[idx]["end_time"]
                    slot["section"] = timeline_slots[idx]["section"]
                    slot["lyrics"] = timeline_slots[idx]["lyrics"]
                    
        return result_json
    except json.JSONDecodeError as e:
        print(f"Failed to parse Gemini response as JSON: {e}")
        print("Raw response:", response.text)
        return None
    finally:
        # アップロードした音声ファイルをクリーンアップして削除
        if audio_file_ref:
            try:
                print(f"[INFO] Deleting audio file from Gemini API: {audio_file_ref.name}")
                genai.delete_file(audio_file_ref.name)
            except Exception as e:
                print(f"[WARNING] Failed to delete file from Gemini API: {e}")

def refine_image_prompt(current_prompt, user_feedback, api_key=None):
    """
    現在の画像プロンプトとユーザーからの日本語のフィードバックを元に、
    より高品質なImagen用の英語プロンプトを生成して返す。
    """
    if api_key:
        genai.configure(api_key=api_key)
    elif os.environ.get("GEMINI_API_KEY"):
        genai.configure(api_key=os.environ.get("GEMINI_API_KEY"))
    else:
        raise ValueError("Gemini API key is not configured.")

    model = genai.GenerativeModel('gemini-2.5-flash')
    
    prompt = f"""
    あなたは画像生成AI (Imagen 3/4) 用のプロンプトエンジニアです。
    ユーザーが作成しているミュージックビデオ(MV)の1カットの画像プロンプトを、ユーザーからの修正指示に基づいてブラッシュアップしてください。

    【現在のプロンプト (英語)】
    {current_prompt}

    【ユーザーからの日本語の修正・追加指示】
    {user_feedback}

    【ブラッシュアップの条件】
    - アスペクト比は16:9を意識してください。
    - 修正指示をプロンプト内に自然に、かつ強力に反映させ、画質やディテール（ライティング、カメラアングル、質感、スタイルなど）を向上させる単語を追加して、詳細な英語の画像生成プロンプトを作成してください。
    - 応答には、ブラッシュアップした「新しい英語プロンプト」のみを出力してください。他の説明文や導入文、コードブロック用のバッククォートなどは一切不要です。
    """

    response = model.generate_content(prompt)
    return response.text.strip().replace('`', '')

def export_project_zip(project_name, timeline_data, bpm, duration, audio_path, images_dir):
    """
    タイムラインJSON、音源ファイル、および生成された画像フォルダをまとめて1つのZIPファイルのバイナリとして返す。
    """
    temp_dir = tempfile.mkdtemp()
    try:
        # 1. JSONの作成
        export_data = {
            "project_name": project_name,
            "bpm": bpm,
            "duration": duration,
            "timeline": timeline_data
        }
        json_path = os.path.join(temp_dir, "project.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
            
        # 2. 音源のコピー
        if audio_path and os.path.exists(audio_path):
            audio_ext = os.path.splitext(audio_path)[1]
            shutil.copy(audio_path, os.path.join(temp_dir, f"audio{audio_ext}"))
            
        # 3. 画像のコピー
        zip_images_dir = os.path.join(temp_dir, "images")
        os.makedirs(zip_images_dir, exist_ok=True)
        if images_dir and os.path.exists(images_dir):
            for filename in os.listdir(images_dir):
                if filename.endswith(".png") and not filename.startswith("temp_"):
                    shutil.copy(os.path.join(images_dir, filename), os.path.join(zip_images_dir, filename))
                    
        # 4. ZIP化
        zip_path = os.path.join(temp_dir, f"{project_name}_package")
        shutil.make_archive(zip_path, 'zip', temp_dir)
        
        # 5. バイナリの読み込み
        with open(zip_path + ".zip", "rb") as f:
            zip_bytes = f.read()
            
        return zip_bytes
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

def import_project_zip(zip_file, extract_dir, images_target_dir):
    """
    アップロードされたZIPを展開し、プロジェクトデータを復元する。
    """
    os.makedirs(extract_dir, exist_ok=True)
    os.makedirs(images_target_dir, exist_ok=True)
    
    # メモリ上のZIPを展開
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        zip_ref.extractall(extract_dir)
        
    # project.json の読み込み
    json_path = os.path.join(extract_dir, "project.json")
    if not os.path.exists(json_path):
        raise ValueError("ZIP内に project.json が見つかりません。")
        
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    # 音源の復元とパス設定
    audio_path = None
    for filename in os.listdir(extract_dir):
        if filename.startswith("audio."):
            target_audio_path = os.path.join(os.path.dirname(images_target_dir), filename)
            shutil.copy(os.path.join(extract_dir, filename), target_audio_path)
            audio_path = target_audio_path
            break
            
    # 画像の復元
    zip_images_dir = os.path.join(extract_dir, "images")
    if os.path.exists(zip_images_dir):
        for filename in os.listdir(zip_images_dir):
            if filename.endswith(".png"):
                shutil.copy(os.path.join(zip_images_dir, filename), os.path.join(images_target_dir, filename))
                
    shutil.rmtree(extract_dir, ignore_errors=True)
    
    return {
        "project_name": data.get("project_name", "imported_project"),
        "bpm": data.get("bpm", 120.0),
        "duration": data.get("duration", 180.0),
        "timeline": data.get("timeline", []),
        "audio_path": audio_path
    }

def export_project_xlsx(project_name, timeline_data, bpm, duration, core_concept="", global_rules="", 
                        first_bar_offset=0.0, time_signature="4/4", split_unit="4小節ごと", 
                        visual_style="特なし", aspect_ratio="16:9"):
    """
    プロジェクトデータをPandasとopenpyxlを使ってスタイリングされたExcelファイル（.xlsx）のバイナリデータとして返す。
    """
    # 1. Summary シート用 DataFrame
    summary_data = {
        "項目": [
            "提案タイトル",
            "BPM",
            "曲の長さ (秒)",
            "世界観・コンセプト",
            "全体演出ルール",
            "1小節目開始時間 (秒)",
            "拍子",
            "分割粒度",
            "ビジュアルスタイル",
            "アスペクト比"
        ],
        "内容": [
            project_name,
            bpm,
            duration,
            core_concept,
            global_rules,
            first_bar_offset,
            time_signature,
            split_unit,
            visual_style,
            aspect_ratio
        ]
    }
    df_summary = pd.DataFrame(summary_data)
    
    # 2. Timeline シート用 DataFrame
    timeline_rows = []
    for i, item in enumerate(timeline_data):
        timeline_rows.append({
            "Cut": i + 1,
            "section": item.get("section", ""),
            "start_time": item.get("start_time", 0.0),
            "end_time": item.get("end_time", 0.0),
            "start_time_ms": f"{int(item.get('start_time', 0.0) // 60):02d}:{item.get('start_time', 0.0) % 60:06.3f}",
            "end_time_ms": f"{int(item.get('end_time', 0.0) // 60):02d}:{item.get('end_time', 0.0) % 60:06.3f}",
            "duration": round(item.get("end_time", 0.0) - item.get("start_time", 0.0), 3),
            "lyrics": item.get("lyrics", ""),
            "description": item.get("description", ""),
            "prompt": item.get("prompt", "")
        })
    df_timeline = pd.DataFrame(timeline_rows)
    
    # メモリ上に出力
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        df_timeline.to_excel(writer, sheet_name="Timeline", index=False)
        
        workbook = writer.book
        
        # --- Summary シートの装飾 ---
        ws_summary = workbook["Summary"]
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # 濃い青
        header_font = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Meiryo UI", size=10)
        
        for col in range(1, 3):
            cell = ws_summary.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row in range(2, len(df_summary) + 2):
            for col in range(1, 3):
                cell = ws_summary.cell(row=row, column=col)
                cell.font = cell_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 60
        
        # --- Timeline シートの装飾 ---
        ws_timeline = workbook["Timeline"]
        
        for col in range(1, len(df_timeline.columns) + 1):
            cell = ws_timeline.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for row in range(2, len(df_timeline) + 2):
            for col in range(1, len(df_timeline.columns) + 1):
                cell = ws_timeline.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = thin_border
                
                col_name = df_timeline.columns[col-1]
                if col_name in ["Cut", "start_time", "end_time", "start_time_ms", "end_time_ms", "duration", "section"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    
        for col in ws_timeline.columns:
            col_letter = get_column_letter(col[0].column)
            col_name = df_timeline.columns[col[0].column - 1]
            
            if col_name in ["lyrics", "description", "prompt"]:
                ws_timeline.column_dimensions[col_letter].width = 40
            elif col_name in ["start_time_ms", "end_time_ms"]:
                ws_timeline.column_dimensions[col_letter].width = 15
            elif col_name in ["Cut", "start_time", "end_time", "duration", "section"]:
                ws_timeline.column_dimensions[col_letter].width = 10
            else:
                ws_timeline.column_dimensions[col_letter].width = 12
                
    output.seek(0)
    return output.getvalue()
