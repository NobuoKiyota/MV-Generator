import os
import sys
import tkinter as tk
from unittest.mock import patch, MagicMock
import openpyxl
import time

# テスト用にGUIアプリをインポート
from xlsx_generator_gui import XlsxGeneratorApp

def run_gui_scenario_test():
    print("====================================================")
    print(" RUNNING NATIVE GUI SCENARIO TEST ")
    print("====================================================")
    
    root = tk.Tk()
    app = XlsxGeneratorApp(root)
    
    # 1. 初期インプットの流し込み
    app.api_key.set("mock-api-key-12345")
    app.audio_type.set("vocal")
    app.toggle_lyrics_state()
    
    # 歌詞エリアへの挿入
    app.txt_lyrics.insert("1.0", "[Intro]\nHello World\n[Chorus]\nFly high to the sky")
    app.raw_idea.set("Cyberpunk aesthetic, neon blue tone")
    
    # 2. STEP 1: コンセプト生成のシミュレーション (API呼び出し部をモック)
    mock_concept = {
        "title": "Neon Flight",
        "concept": "A futuristic city under neon lights with raining code.",
        "characters": "A female hacker with silver hair and glowing blue eyes, wearing a black cyber-jacket.",
        "rules": "Neon blue and yellow color palette, fast camera cuts, cyberpunk aesthetic."
    }
    
    print("[TEST] Simulating STEP 1: Concept Generation...")
    
    # API呼び出しのモック
    with patch("xlsx_generator_gui.generate_concept_settings", return_value=mock_concept) as mock_api_1:
        # 非同期スレッドが本来走るが、テストのため直接プロセス処理を叩いて動作を確認する
        # (スレッド内部で発生するUI書き換えロジックがエラーを吐かないかを確認)
        app.generate_concept_process()
        
        # 呼び出されたか確認
        mock_api_1.assert_called_once()
        
    # Tkinterのafterキューを実行するためにアップデートを強制呼び出し
    root.update()
    
    # UIにデータが正しく反映されたか検証
    title = app.concept_title.get().strip()
    concept_text = app.txt_res_concept.get("1.0", tk.END).strip()
    char_text = app.txt_res_char.get("1.0", tk.END).strip()
    rules_text = app.txt_res_rules.get("1.0", tk.END).strip()
    
    assert title == "Neon Flight", f"Expected 'Neon Flight', got '{title}'"
    assert "futuristic city" in concept_text, "Concept text was not updated correctly."
    assert "female hacker" in char_text, "Character text was not updated correctly."
    assert "Neon blue" in rules_text, "Rules text was not updated correctly."
    
    print("[TEST] STEP 1 check passed. (Worldview definitions verified)")

    # 2.5. テスト画像生成 (Preview) のシミュレーション
    print("[TEST] Simulating Test Image Preview Generation...")
    
    mock_test_prompt = "A high detailed futuristic neon city, cyberpunk style, 16:9 aspect ratio."
    
    # テスト画像生成時に Pillow で実際にダミー画像ファイルを作成して保存する
    def side_effect_generate_image(prompt, output_path, api_key=None):
        from PIL import Image
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        img = Image.new('RGB', (240, 135), color=(0, 0, 255))
        img.save(output_path)
        return True, "LOCAL_SUCCESS"

    with patch("xlsx_generator_gui.generate_test_image_prompt", return_value=mock_test_prompt) as mock_prompt_gen, \
         patch("xlsx_generator_gui.generate_imagen_image", side_effect=side_effect_generate_image) as mock_img_gen:
         
        app.generate_test_image_process()
        
        mock_prompt_gen.assert_called_once()
        mock_img_gen.assert_called_once()

    root.update()
    
    # プレビュー画像がGUIに正しく読み込まれたか検証
    assert app.preview_photo is not None, "Preview photo was not loaded in App."
    # テンポラリプレビューファイルのクリーンアップ
    temp_preview_file = os.path.join("assets", "temp_preview.png")
    if os.path.exists(temp_preview_file):
        os.remove(temp_preview_file)
        
    print("[TEST] Test Image Preview check passed.")

    # 3. STEP 2: タイムライン & Excel生成のシミュレーション
    mock_timeline = {
        "bpm": 120.0,
        "duration": 180.0,
        "summary": {
            "title": "Neon Flight",
            "concept": "A futuristic city...",
            "characters": "A female hacker...",
            "rules": "Neon blue..."
        },
        "timeline": [
            {
                "section": "Intro",
                "start_time": 0.0,
                "end_time": 5.0,
                "lyrics": "",
                "description": "Establish shot of the neon city.",
                "prompt": "Neon city, cyberpunk aesthetic"
            },
            {
                "section": "Chorus",
                "start_time": 5.0,
                "end_time": 15.0,
                "lyrics": "Fly high to the sky",
                "description": "A female hacker looks up at the sky.",
                "prompt": "Female hacker with silver hair, neon city background"
            }
        ]
    }
    
    test_excel_path = os.path.join(os.path.dirname(__file__), "test_output_timeline.xlsx")
    if os.path.exists(test_excel_path):
        os.remove(test_excel_path)
        
    print("[TEST] Simulating STEP 2: Timeline Generation...")
    
    # タイムライン生成APIのモック、およびファイル保存先ダイアログのモック
    with patch("xlsx_generator_gui.generate_timeline_from_concept", return_value=mock_timeline) as mock_api_2, \
         patch("tkinter.filedialog.asksaveasfilename", return_value=test_excel_path) as mock_dialog:
         
        app.generate_timeline_process()
        
        mock_api_2.assert_called_once()
        mock_dialog.assert_called_once()
        
    # Tkinterのafterキューを実行するためにアップデートを強制呼び出し
    root.update()
    
    # Excelが実際に正しく生成されたか検証
    assert os.path.exists(test_excel_path), "Excel file was not generated."
    
    # openpyxlでロードして内容を検証
    wb = openpyxl.load_workbook(test_excel_path)
    assert "Summary" in wb.sheetnames, "Summary sheet is missing."
    assert "Timeline" in wb.sheetnames, "Timeline sheet is missing."
    
    ws_summary = wb["Summary"]
    # A1=項目, B1=内容
    # B1(タイトル)=Neon Flight (A列は項目名、B列は内容)
    assert ws_summary["B2"].value == "Neon Flight", "Project title in Excel is incorrect."
    
    ws_timeline = wb["Timeline"]
    # 1行目はヘッダー、2行目はカット1、3行目はカット2
    assert ws_timeline["B2"].value == "Intro", "Timeline section is incorrect."
    assert ws_timeline["H3"].value == "Fly high to the sky", "Lyrics content in Timeline sheet is incorrect."
    
    wb.close()
    
    # テスト用ファイルを削除
    if os.path.exists(test_excel_path):
        os.remove(test_excel_path)
        
    print("[TEST] STEP 2 check passed.")
    
    # Tkインフラの破棄
    root.destroy()
    
    print("====================================================")
    print(" ALL SCENARIO TESTS PASSED SUCCESSFULLY! ")
    print("====================================================")
    sys.exit(0)

if __name__ == "__main__":
    try:
        run_gui_scenario_test()
    except Exception as e:
        print(f"[ERROR] TEST FAILED: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
